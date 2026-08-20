"""Своды по архиву начислений: по домам и по абонентам внутри дома.

Форма повторяет бумажный отчёт бухгалтерии: двухъярусная шапка
«Сальдо нач. месяца | Начислено | Оплачено | Перевод | Пеня | Возврат |
Сальдо кон. месяца | Просроченная задолженность | Вид ком. услуг».

Свод считается агрегацией по строкам `PeriodSnapshot`, а не хранится отдельной
таблицей: итог по компании и расшифровка по жильцам берутся из одних и тех же
строк и поэтому не могут разойтись.

Набор колонок описан здесь один раз — им пользуются и HTML-предпросмотр, и PDF,
и выгрузка в Excel, чтобы три формы одного отчёта не разъезжались.
"""

import re

from django.db.models import Count, Sum

from .models import MONTHS, DEFAULT_SERVICE_TYPE, House, PeriodSnapshot


# Денежные графы: поле модели -> заголовок нижнего яруса шапки.
# Порядок здесь и есть порядок колонок в отчёте.
COLUMN_GROUPS = (
    ('Сальдо нач. месяца', (('opening_debit', 'Дебет'), ('opening_credit', 'Кредит'))),
    ('Начислено', (('charged_ku', 'КУ'), ('charged_fee', '3%'))),
    ('Оплачено', (('paid_ku', 'КУ'), ('paid_fee', '3%'))),
    ('Перевод', (('transfer', ''),)),
    ('Пеня', (('penalty', ''),)),
    ('Возврат', (('refund', ''),)),
    ('Сальдо кон. месяца', (('closing_debit', 'Дебет'), ('closing_credit', 'Кредит'))),
    ('Просроченная задолжен.', (('overdue', ''),)),
)

MONEY_FIELDS = tuple(field for _, columns in COLUMN_GROUPS for field, _ in columns)

MODE_HOUSES = 'houses'
MODE_SUBSCRIBERS = 'subscribers'

# Колонки слева от денежных — разные у свода по домам и у расшифровки.
LEADING_COLUMNS = {
    MODE_HOUSES: (('label', 'Адрес'), ('accounts', 'Лиц. счетов')),
    MODE_SUBSCRIBERS: (('ls', 'Лицевой счет'), ('fio', 'ФИО'), ('apartment', 'Кв.')),
}

MODE_TITLES = {
    MODE_HOUSES: 'Свод по домам',
    MODE_SUBSCRIBERS: 'Расшифровка по абонентам',
}


def month_name(month):
    return MONTHS.get(month, str(month))


def period_label(year, month):
    return f'{month_name(month)} {year} г.'


def available_periods():
    """Периоды, за которые архив уже есть, — от новых к старым."""
    return list(
        PeriodSnapshot.objects
        .values_list('year', 'month')
        .distinct()
        .order_by('-year', '-month')
    )


def latest_period():
    periods = available_periods()
    return periods[0] if periods else (None, None)


def available_houses(year=None, month=None):
    """Дома, по которым есть архив за период (для выпадающего фильтра)."""
    queryset = PeriodSnapshot.objects.filter(house__isnull=False)
    if year and month:
        queryset = queryset.filter(year=year, month=month)
    return House.objects.filter(pk__in=queryset.values('house')).order_by(
        'street', 'number_order', 'number')


def snapshots(year, month, house=None, only_debtors=False):
    queryset = PeriodSnapshot.objects.filter(year=year, month=month)
    if house is not None:
        queryset = queryset.filter(house=house)
    if only_debtors:
        # Должник — тот, у кого на конец месяца остался долг, а не только тот,
        # у кого просрочка: домкому нужны и те, кто не заплатил за этот месяц.
        queryset = queryset.filter(closing_debit__gt=0)
    return queryset


def _sums():
    return {field: Sum(field) for field in MONEY_FIELDS}


def totals(queryset):
    """Итоговая строка: суммы по всем денежным графам."""
    aggregated = queryset.aggregate(accounts=Count('pk'), **_sums())
    row = {'accounts': aggregated['accounts'] or 0}
    row.update({field: round(aggregated[field] or 0.0, 2) for field in MONEY_FIELDS})
    return row


def house_rows(queryset):
    """Строки свода: одна строка на дом, как в бумажной форме."""
    aggregated = (
        queryset
        .values('house', 'house__street', 'house__number', 'house__number_order',
                'house__service_type')
        .annotate(accounts=Count('pk'), **_sums())
        .order_by('house__street', 'house__number_order', 'house__number')
    )

    rows = []
    for item in aggregated:
        street = item['house__street']
        number = item['house__number']
        if item['house'] is None:
            # Абоненты с неразобранным адресом. Прятать их нельзя: тогда сумма
            # строк перестанет сходиться с итогом.
            label = 'Без дома'
        else:
            label = f'{street} дом {number}' if street and number else (street or number)

        row = {
            'house_id': item['house'],
            'label': label,
            'accounts': item['accounts'],
            'service_type': item['house__service_type'] or DEFAULT_SERVICE_TYPE,
        }
        row.update({field: round(item[field] or 0.0, 2) for field in MONEY_FIELDS})
        rows.append(row)
    return rows


def apartment_key(value):
    """Естественный порядок квартир: 2 перед 10, «12а» после «12»."""
    match = re.match(r'(\d+)(.*)', str(value or '').strip())
    if match:
        return (0, int(match.group(1)), match.group(2))
    return (1, 0, str(value or ''))


def subscriber_rows(queryset):
    """Строки расшифровки: один абонент — одна строка, по порядку квартир."""
    rows = []
    for snapshot in queryset.select_related('house'):
        row = {
            'pk': snapshot.pk,
            'ls': snapshot.ls,
            'fio': snapshot.fio,
            'apartment': snapshot.apartment,
            'label': snapshot.address,
            'accounts': 1,
            'service_type': snapshot.service_type,
        }
        row.update({field: round(getattr(snapshot, field) or 0.0, 2)
                    for field in MONEY_FIELDS})
        rows.append(row)
    rows.sort(key=lambda row: apartment_key(row['apartment']))
    return rows


def build_report(year, month, house=None, mode=MODE_HOUSES, only_debtors=False):
    """Готовый контекст отчёта — общий для HTML, PDF и Excel."""
    mode = mode if mode in LEADING_COLUMNS else MODE_HOUSES
    queryset = snapshots(year, month, house=house, only_debtors=only_debtors)
    rows = subscriber_rows(queryset) if mode == MODE_SUBSCRIBERS else house_rows(queryset)

    return {
        'mode': mode,
        'title': MODE_TITLES[mode],
        'year': year,
        'month': month,
        'period': period_label(year, month),
        'house': house,
        'only_debtors': only_debtors,
        'rows': rows,
        'totals': totals(queryset),
        'column_groups': COLUMN_GROUPS,
        'leading_columns': LEADING_COLUMNS[mode],
        'money_fields': MONEY_FIELDS,
        'service_type': house.service_type if house is not None else DEFAULT_SERVICE_TYPE,
        'is_empty': not rows,
    }


def flat_columns(report):
    """Плоский список колонок (ключ, заголовок) — для выгрузки в Excel."""
    columns = list(report['leading_columns'])
    for group, group_columns in COLUMN_GROUPS:
        for field, subtitle in group_columns:
            columns.append((field, f'{group} {subtitle}'.strip()))
    columns.append(('service_type', 'Вид ком. услуг'))
    return columns


def report_filename(report, extension):
    parts = ['svod', str(report['year']), f"{report['month']:02d}"]
    if report['house'] is not None:
        parts.append(f"dom-{report['house'].pk}")
    if report['mode'] == MODE_SUBSCRIBERS:
        parts.append('abonenty')
    return '_'.join(parts) + f'.{extension}'


def workbook_for(report):
    """Тот же свод в Excel — с двухъярусной шапкой, как на бумаге."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, Side
    from openpyxl.utils import get_column_letter

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = f"{report['year']}-{report['month']:02d}"

    thin = Side(style='thin')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    centered = Alignment(horizontal='center', vertical='center', wrap_text=True)

    subtitle = report['period']
    if report['house'] is not None:
        subtitle = f"{report['house']} — {subtitle}"
    sheet.append([f"{report['title']}. {subtitle}"])
    sheet['A1'].font = Font(bold=True, size=12)
    sheet.append([])

    header_top, header_bottom = 3, 4
    column = 1

    def put(row, col, value):
        cell = sheet.cell(row=row, column=col, value=value)
        cell.font = Font(bold=True)
        cell.alignment = centered
        cell.border = border
        return cell

    for _, title in report['leading_columns']:
        put(header_top, column, title)
        put(header_bottom, column, None)
        sheet.merge_cells(start_row=header_top, start_column=column,
                          end_row=header_bottom, end_column=column)
        column += 1

    for group, group_columns in COLUMN_GROUPS:
        put(header_top, column, group)
        if len(group_columns) > 1:
            sheet.merge_cells(start_row=header_top, start_column=column,
                              end_row=header_top, end_column=column + len(group_columns) - 1)
            for offset, (_, sub) in enumerate(group_columns):
                put(header_bottom, column + offset, sub)
        else:
            put(header_bottom, column, None)
            sheet.merge_cells(start_row=header_top, start_column=column,
                              end_row=header_bottom, end_column=column)
        column += len(group_columns)

    put(header_top, column, 'Вид ком. услуг')
    put(header_bottom, column, None)
    sheet.merge_cells(start_row=header_top, start_column=column,
                      end_row=header_bottom, end_column=column)
    last_column = column

    columns = flat_columns(report)
    for row in report['rows'] + [_totals_row(report)]:
        values = [row.get(key, '') for key, _ in columns]
        sheet.append(values)
        for index in range(1, last_column + 1):
            cell = sheet.cell(row=sheet.max_row, column=index)
            cell.border = border
            if isinstance(cell.value, float):
                cell.number_format = '#,##0.00'
        if row.get('is_total'):
            for index in range(1, last_column + 1):
                sheet.cell(row=sheet.max_row, column=index).font = Font(bold=True)

    widths = {1: 34} if report['mode'] == MODE_HOUSES else {1: 16, 2: 32, 3: 8}
    for index in range(1, last_column + 1):
        sheet.column_dimensions[get_column_letter(index)].width = widths.get(index, 13)
    sheet.freeze_panes = sheet.cell(row=header_bottom + 1, column=1)

    return workbook


def _totals_row(report):
    """Строка «ИТОГО» в том же виде, что и обычные строки таблицы."""
    row = dict(report['totals'])
    row.update({
        'is_total': True,
        'label': 'ИТОГО:',
        'ls': 'ИТОГО:',
        'fio': '',
        'apartment': '',
        'service_type': report['service_type'],
    })
    return row


def format_money(value):
    """Сумма для печати: «1 234,56». Ноль — пустая клетка, как на бумаге."""
    if value in (None, ''):
        return ''
    value = round(float(value), 2)
    if abs(value) < 0.005:
        return ''
    return f'{value:,.2f}'.replace(',', ' ').replace('.', ',')


def _cell(value, key):
    if key in MONEY_FIELDS:
        return {'value': format_money(value), 'align': 'num'}
    if key == 'accounts':
        return {'value': value or '', 'align': 'num'}
    return {'value': value if value not in (None, '') else '', 'align': 'txt'}


def table_rows(report):
    """Строки таблицы для HTML и PDF, последняя — «ИТОГО»."""
    columns = flat_columns(report)
    rows = [
        {'cells': [_cell(row.get(key, ''), key) for key, _ in columns], 'is_total': False}
        for row in report['rows']
    ]
    total = _totals_row(report)
    rows.append({
        'cells': [_cell(total.get(key, ''), key) for key, _ in columns],
        'is_total': True,
    })
    return rows


def workbook_for_snapshots(queryset):
    """Плоская выгрузка произвольной выборки архива.

    В отличие от `workbook_for` здесь нет двухъярусной шапки и итогов: это не
    бланк отчёта, а таблица для разбора вручную, куда может попасть сразу
    несколько месяцев.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Архив'

    headers = ['Год', 'Месяц', 'Лицевой счет', 'ФИО', 'Адрес', 'Кв.']
    fields = []
    for group, group_columns in COLUMN_GROUPS:
        for field, subtitle in group_columns:
            headers.append(f'{group} {subtitle}'.strip())
            fields.append(field)
    headers.append('Вид ком. услуг')

    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True)

    for snapshot in queryset.select_related('house').order_by('-year', '-month', 'address'):
        row = [snapshot.year, month_name(snapshot.month), snapshot.ls, snapshot.fio,
               snapshot.address, snapshot.apartment]
        row += [round(getattr(snapshot, field) or 0.0, 2) for field in fields]
        row.append(snapshot.service_type)
        sheet.append(row)
        for index in range(7, 7 + len(fields)):
            sheet.cell(row=sheet.max_row, column=index).number_format = '#,##0.00'

    widths = {1: 7, 2: 12, 3: 16, 4: 32, 5: 34, 6: 7}
    for index in range(1, len(headers) + 1):
        sheet.column_dimensions[get_column_letter(index)].width = widths.get(index, 13)
    sheet.freeze_panes = 'A2'

    return workbook
